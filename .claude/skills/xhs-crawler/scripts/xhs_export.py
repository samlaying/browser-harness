#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""小红书 Excel 导出 — 含图片下载、单 Sheet 帖子-评论结构

用法：
    python3 xhs_export.py <json_dir_or_file> [output.xlsx]

输入：xhs_crawl.py 输出的 JSON 文件或包含多个 JSON 的目录
输出：单 Sheet Excel，帖子标题行（彩色）+ 评论行缩进，图片嵌入

依赖：pip install openpyxl
"""

import json, os, sys, datetime, glob, urllib.request
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# ── 颜色 ─────────────────────────────────────────────────

POST_COLORS = [
    "D6EAF8", "D5F5E3", "FCF3CF", "FADBD8", "E8DAEF",
    "D4E6F1", "ABEBC6", "F9E79F", "F5B7B1", "D2B4DE",
    "AED6F1", "82E0AA", "F8C471", "F1948A", "BB8FCE",
    "85C1E9", "58D68D", "F0B27A", "EC7063", "AF7AC5",
]

THIN_BORDER = Border(
    bottom=Side(style='thin', color='CCCCCC'),
)

# ── 图片下载 ──────────────────────────────────────────────

def download_image(url, fpath):
    if os.path.exists(fpath) and os.path.getsize(fpath) > 0:
        return True
    try:
        req = urllib.request.Request(url, headers={
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)',
            'Referer': 'https://www.xiaohongshu.com/',
        })
        with urllib.request.urlopen(req, timeout=15) as resp:
            with open(fpath, 'wb') as f:
                f.write(resp.read())
        return True
    except Exception as e:
        print(f"  下载失败: {e}")
        return False

# ── 主逻辑 ────────────────────────────────────────────────

def load_notes(path):
    """加载 JSON 文件或目录下所有 JSON（只收笔记字典，跳过 notes_order.json 等非笔记文件）"""
    if os.path.isfile(path):
        with open(path, encoding='utf-8') as f:
            return [json.load(f)]
    files = sorted(glob.glob(os.path.join(path, '*.json')))
    notes = []
    for fp in files:
        if os.path.basename(fp) == 'notes_order.json':
            continue
        try:
            with open(fp, encoding='utf-8') as f:
                d = json.load(f)
            if isinstance(d, dict):
                notes.append(d)
        except: pass
    return notes

def build_excel(notes, output_path, img_dir):
    wb = Workbook()
    ws = wb.active
    ws.title = "小红书数据"

    # 列定义
    # A:序号 B:层级 C:昵称 D:评论内容 E:回复→ F:时间 G:IP H:赞 I:帖子图片
    headers = ['序号', '层级', '昵称', '评论内容', '回复→', '时间', 'IP', '赞', '帖子图片']
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 55
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 8
    ws.column_dimensions['H'].width = 6
    ws.column_dimensions['I'].width = 18

    row = 1
    all_dl_map = {}  # url -> local path
    comment_seq = 0

    for note_idx, note in enumerate(notes):
        meta = note.get('meta', {})
        comments = note.get('comments', [])
        color_hex = POST_COLORS[note_idx % len(POST_COLORS)]
        fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

        title = meta.get('title', '未知标题')
        bar = meta.get('barText', '')

        # ── 帖子标题行 ──
        ws.cell(row=row, column=1, value=f"{note_idx+1}.")
        ws.cell(row=row, column=2, value="帖子")
        ws.cell(row=row, column=3, value=title[:30])
        ws.cell(row=row, column=4, value=meta.get('desc', '')[:200])
        ws.cell(row=row, column=5, value=bar)
        ws.cell(row=row, column=9, value=f"图片×{len(meta.get('noteImgs', []))}")

        # 标题行样式
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.font = Font(bold=True, size=11)
        ws.row_dimensions[row].height = 30

        # 帖子图片：紧贴标题行右侧（I 列）
        noteImgs = meta.get('noteImgs', [])
        post_img_row = row
        for img_idx, img_url in enumerate(noteImgs):
            img_row = post_img_row + img_idx
            fpath = os.path.join(img_dir, f'post_{note_idx}_{img_idx}.jpg')
            download_image(img_url, fpath)
            all_dl_map[img_url] = fpath

            if img_idx > 0:
                # 多图时每张占新行
                ws.row_dimensions[img_row].height = 120 * 0.75
                for col in range(1, 9):
                    ws.cell(row=img_row, column=col).fill = fill

            if os.path.exists(fpath) and os.path.getsize(fpath) > 0:
                try:
                    img = XlImage(fpath)
                    img.width = 120; img.height = 120
                    img.anchor = OneCellAnchor(
                        _from=AnchorMarker(col=8, colOff=0, row=img_row-1, rowOff=0),
                        ext=XDRPositiveSize2D(pixels_to_EMU(120), pixels_to_EMU(120)),
                    )
                    ws.add_image(img)
                except: pass

        row += 1

        # ── 评论行 ──
        if not comments:
            ws.cell(row=row, column=2, value="无评论")
            ws.cell(row=row, column=3, value="—")
            for col in range(1, 10):
                ws.cell(row=row, column=col).fill = fill
            ws.row_dimensions[row].height = 22
            row += 1
        else:
            # 构建线程关系
            thread_id = 0; current_thread = 0
            for c in comments:
                if c['lvl'] == 1:
                    thread_id += 1; current_thread = thread_id
                    c['thread_id'] = thread_id; c['reply_to'] = ''
                else:
                    c['thread_id'] = current_thread; c['reply_to'] = ''
            for i, c in enumerate(comments):
                if c['lvl'] == 2:
                    for j in range(i-1, -1, -1):
                        if comments[j]['nick'] != c['nick']:
                            c['reply_to'] = comments[j]['nick']; break

            for c in comments:
                comment_seq += 1
                lvl = '评论' if c['lvl'] == 1 else '↳回复'
                ws.cell(row=row, column=1, value=comment_seq)
                ws.cell(row=row, column=2, value=lvl)
                ws.cell(row=row, column=3, value=c.get('nick', ''))
                ws.cell(row=row, column=4, value=c.get('content', ''))
                ws.cell(row=row, column=5, value=c.get('reply_to', ''))
                ws.cell(row=row, column=6, value=c.get('date', ''))
                ws.cell(row=row, column=7, value=c.get('ip', ''))
                ws.cell(row=row, column=8, value=c.get('likes', '0'))

                for col in range(1, 10):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = fill
                    cell.border = THIN_BORDER
                    if c['lvl'] == 2:
                        cell.font = Font(color="555555", size=10)

                ws.row_dimensions[row].height = 22

                # 评论图片
                cmt_imgs = c.get('imgs', [])
                if cmt_imgs:
                    ws.row_dimensions[row].height = 60 * 0.75
                    for img_url in cmt_imgs[:2]:
                        fpath = os.path.join(img_dir, f'cmt_{comment_seq}.jpg')
                        download_image(img_url, fpath)
                        if os.path.exists(fpath) and os.path.getsize(fpath) > 0:
                            try:
                                img = XlImage(fpath)
                                img.width = 60; img.height = 60
                                img.anchor = OneCellAnchor(
                                    _from=AnchorMarker(col=3, colOff=pixels_to_EMU(150), row=row-1, rowOff=0),
                                    ext=XDRPositiveSize2D(pixels_to_EMU(60), pixels_to_EMU(60)),
                                )
                                ws.add_image(img)
                            except: pass

                row += 1

        # 帖子间分隔线
        row += 1

    wb.save(output_path)
    return output_path

# ── 入口 ──────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("用法: python3 xhs_export.py <json_dir_or_file> [output.xlsx]")
        sys.exit(1)

    path = sys.argv[1]
    base_dir = os.path.dirname(path) if os.path.isfile(path) else path
    img_dir = os.path.join(base_dir, 'xhs_images')
    os.makedirs(img_dir, exist_ok=True)

    notes = load_notes(path)
    # 去重（按标题）
    seen_titles = set()
    unique = []
    for n in notes:
        t = n.get('meta', {}).get('title', '')
        if t and t not in seen_titles:
            seen_titles.add(t)
            unique.append(n)
    notes = unique

    print(f"加载 {len(notes)} 篇笔记")

    if len(sys.argv) >= 3:
        out = sys.argv[2]
    else:
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        out = os.path.join(base_dir, f'xhs_批量_{ts}.xlsx')

    build_excel(notes, out, img_dir)
    total_comments = sum(len(n.get('comments', [])) for n in notes)
    print(f"✅ {len(notes)} 篇帖子, {total_comments} 条评论 → {out}")

if __name__ == '__main__':
    main()
