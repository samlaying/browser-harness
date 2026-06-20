#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""把 AI 巡检结果导出为 Markdown 摘要 + Excel。

用法：
    python3 scripts/x_ai_export.py [ai_hits.json所在目录]

默认目录：xai_data/<YYYYMMDD>_scan（取最新的 xai_data/*_scan）。
依赖：openpyxl（pip install openpyxl）。
"""

import os, sys, json, glob
from datetime import datetime

def find_dir(arg):
    if arg and os.path.isdir(arg):
        return arg
    base = "xai_data"
    cands = sorted(glob.glob(os.path.join(base, "*_scan")), reverse=True)
    return cands[0] if cands else base

def main():
    d = find_dir(sys.argv[1] if len(sys.argv) > 1 else None)
    hits_path = os.path.join(d, "ai_hits.json")
    all_path = os.path.join(d, "all_tweets.json")
    if not os.path.exists(hits_path):
        print(f"✗ 找不到 {hits_path}，先跑 x_ai_scan.py"); return

    hits = json.load(open(hits_path, encoding="utf-8"))
    all_tweets = json.load(open(all_path, encoding="utf-8")) if os.path.exists(all_path) else []
    stamp = datetime.now().strftime("%Y-%m-%d %H:%M")

    # ── Markdown 摘要 ──────────────────────────────────────
    lines = []
    lines.append(f"# X AI 巡检报告\n")
    lines.append(f"- 生成时间：{stamp}")
    lines.append(f"- 数据目录：`{d}`")
    lines.append(f"- 扫描推文总数：{len(all_tweets)}")
    lines.append(f"- AI 相关命中：{len(hits)}\n")

    by_cat = {"project": [], "trend": [], "opinion": [], "general": []}
    for t in hits:
        by_cat.setdefault(t.get("category", "general"), []).append(t)

    titles = {
        "project": "🚀 AI 项目 / 产品",
        "trend": "📈 AI 趋势 / 方向",
        "opinion": "💬 AI 观点 / 看法",
        "general": "📌 其它 AI 相关",
    }
    for cat in ["project", "trend", "opinion", "general"]:
        items = by_cat.get(cat, [])
        if not items:
            continue
        lines.append(f"## {titles.get(cat, cat)} （{len(items)}）\n")
        for t in items:
            txt = (t.get("text") or "").replace("\n", " ").strip()
            terms = ", ".join(t.get("ai_terms", [])) or "-"
            eng = f"❤️{t.get('likes',0)} 🔁{t.get('reposts',0)} 💬{t.get('replies',0)}"
            lines.append(f"### @{t.get('handle','?')}  ·  {eng}")
            lines.append(f"> {txt}\n")
            lines.append(f"- 命中词：{terms}　·  [原文]({t.get('url','')})\n")
    md_path = os.path.join(d, "ai_digest.md")
    open(md_path, "w", encoding="utf-8").write("\n".join(lines))
    print(f"✓ Markdown: {md_path}")

    # ── Excel ─────────────────────────────────────────────
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = "AI 命中"
        headers = ["分类", "作者", "Handle", "正文", "点赞", "转发", "回复", "浏览", "命中词", "时间", "链接"]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="305496")
            c.alignment = Alignment(vertical="center")
        cat_color = {"project":"E2EFDA","trend":"DDEBF7","opinion":"FFF2CC","general":"F2F2F2"}
        for t in hits:
            ws.append([
                t.get("category",""), t.get("name",""), t.get("handle",""),
                (t.get("text") or "").replace("\n"," "),
                t.get("likes",0), t.get("reposts",0), t.get("replies",0), t.get("views",0),
                ", ".join(t.get("ai_terms",[])), t.get("dt",""), t.get("url",""),
            ])
        ws.column_dimensions["D"].width = 70
        ws.column_dimensions["J"].width = 20
        ws.column_dimensions["K"].width = 45
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fill = cat_color.get(hits[i-2].get("category"), "FFFFFF")
            for c in row:
                c.fill = PatternFill("solid", fgColor=fill)
                c.alignment = Alignment(vertical="top", wrap_text=("D" == c.column_letter))
        xlsx_path = os.path.join(d, "ai_scan.xlsx")
        wb.save(xlsx_path)
        print(f"✓ Excel:    {xlsx_path}")
    except ImportError:
        print("⚠ 未装 openpyxl，跳过 Excel（pip install openpyxl）")

    print(f"\n命中 {len(hits)} 条 / 共 {len(all_tweets)} 条")

if __name__ == "__main__":
    main()
