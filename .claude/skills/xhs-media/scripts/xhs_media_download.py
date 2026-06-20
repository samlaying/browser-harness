#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""小红书媒体下载器 + 汇总 Excel — xhs-media skill

下载视频(签名 masterUrl + backupUrls 即时下)、原图、头像到 folder-per-note 布局，
并生成一份媒体汇总 Excel。纯 requests/openpyxl，不依赖 browser-harness。

既可被 xhs_media_batch.py import，也可独立运行：
    python3 xhs_media_download.py <json_dir> [output.xlsx]

输入 json_dir：xhs_media_extract.py 输出的 JSON 目录（每个文件一篇笔记）。
"""

import os
import re
import sys
import json
import glob
import time
import queue
import threading
import argparse
import datetime
import urllib.request
from urllib.parse import urlparse, unquote

import requests

# openpyxl 可选——只在生成 Excel 时需要
try:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAVE_OPENPYXL = True
except Exception:
    HAVE_OPENPYXL = False

UA = 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/137.0.0.0 Safari/537.36'
REFERER = 'https://www.xiaohongshu.com/'

POST_COLORS = [
    "D6EAF8", "D5F5E3", "FCF3CF", "FADBD8", "E8DAEF",
    "D4E6F1", "ABEBC6", "F9E79F", "F5B7B1", "D2B4DE",
]

# ── 文件名工具（复制自 redbook-download/xiaohongshu_downloader.py）──────────

def safe_filename(name, max_length=50):
    """清理文件名非法字符"""
    if not name:
        name = "未命名"
    name = re.sub(r'[\\/*?:"<>|]', "_", name)
    name = name.strip()
    if len(name) > max_length:
        name = name[:max_length]
    name = name.strip('. ')
    return name or "未命名"


def get_file_extension(url):
    """根据 URL 推断扩展名"""
    parsed = urlparse(url)
    path = unquote(parsed.path)
    if '.' in path:
        ext = path.split('.')[-1].lower()
        if ext in ('jpg', 'jpeg', 'png', 'webp', 'gif', 'mp4', 'mov', 'avi', 'webm'):
            return f".{ext}"
    low = url.lower()
    if any(x in low for x in ['video', 'mp4', 'mov']):
        return '.mp4'
    return '.jpg'


def _new_session():
    s = requests.Session()
    s.headers.update({
        'User-Agent': UA,
        'Referer': REFERER,
        'Accept': 'image/webp,image/apng,image/*,*/*;q=0.8',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
    })
    return s


# ── 下载 ──────────────────────────────────────────────────────────────────

def download_url(url, dest, session=None, max_retries=3, min_size=1024, timeout=60):
    """下载单个 URL 到 dest。成功(dest 存在且 >min_size)返回 True。"""
    if os.path.exists(dest) and os.path.getsize(dest) > min_size:
        return True
    session = session or _new_session()
    for attempt in range(max_retries):
        try:
            r = session.get(url, stream=True, timeout=timeout, headers={'Referer': REFERER})
            r.raise_for_status()
            tmp = dest + '.part'
            with open(tmp, 'wb') as f:
                for chunk in r.iter_content(8192):
                    if chunk:
                        f.write(chunk)
            if os.path.getsize(tmp) > min_size:
                os.replace(tmp, dest)
                return True
            os.remove(tmp)
        except Exception as e:
            if attempt < max_retries - 1:
                time.sleep(2 ** attempt)
            else:
                print(f"  ❌ 下载失败 {os.path.basename(dest)}: {e}", flush=True)
    return False


def download_video(media, dest_dir, session):
    """视频：先试签名 masterUrl，失败/过期换 backupUrls。返回本地路径或 None。

    masterUrl 是签名 URL（?sign=&t=），会过期——抓到后必须立刻下，这就是当篇立即下载的原因。
    """
    note_id = media.get('noteId', '') or media.get('note_id', '')
    dest = os.path.join(dest_dir, f"视频_{note_id}.mp4")
    if os.path.exists(dest) and os.path.getsize(dest) > 10240:
        return dest
    candidates = []
    if media.get('videoUrl'):
        candidates.append(media['videoUrl'])
    candidates.extend(media.get('videoBackups') or [])
    for cand in candidates:
        if not cand:
            continue
        if download_url(cand, dest, session, max_retries=2, min_size=10240):
            print(f"  🎬 视频: {os.path.basename(dest)} ({os.path.getsize(dest)//1024}KB)", flush=True)
            return dest
    return None


def download_images(media, dest_dir, session):
    """下载所有原图。返回 [本地路径]"""
    out = []
    for i, url in enumerate(media.get('imageUrls') or [], 1):
        ext = get_file_extension(url)
        dest = os.path.join(dest_dir, f"图片_{i:02d}{ext}")
        if download_url(url, dest, session, max_retries=3, min_size=1024):
            out.append(dest)
    if out:
        print(f"  🖼️  图片: {len(out)} 张", flush=True)
    return out


def download_avatar(media, dest_dir, session):
    """下载头像"""
    url = media.get('avatar')
    if not url:
        return None
    nick = safe_filename(media.get('nickname', '用户'), max_length=30)
    dest = os.path.join(dest_dir, f"头像_{nick}{get_file_extension(url)}")
    if download_url(url, dest, session, max_retries=3, min_size=512):
        return dest
    return None


def download_note_media(media, base_dir, index, session=None):
    """下载单篇笔记全部媒体到 {base_dir}/{idx:03d}_{title}_{noteId[:8]}/。

    返回 dict：{note_dir, downloaded_files[], video_path, image_count, has_avatar}
    """
    session = session or _new_session()
    title = safe_filename(media.get('title') or f'笔记_{index}', max_length=40)
    note_id = media.get('noteId') or media.get('note_id') or f'note_{index}'
    note_dir = os.path.join(base_dir, f"{index:03d}_{title}_{note_id[:8]}")
    os.makedirs(note_dir, exist_ok=True)

    downloaded = []
    is_video = (media.get('type') == 'video')

    if is_video:
        v = download_video(media, note_dir, session)
        if v:
            downloaded.append(os.path.basename(v))

    imgs = download_images(media, note_dir, session)
    downloaded.extend(os.path.basename(p) for p in imgs)

    av = download_avatar(media, note_dir, session)
    if av:
        downloaded.append(os.path.basename(av))

    # 笔记信息.json：全部字段 + 下载清单
    info = dict(media)
    info['downloaded_files'] = downloaded
    info['is_video'] = is_video
    with open(os.path.join(note_dir, '笔记信息.json'), 'w', encoding='utf-8') as f:
        json.dump(info, f, ensure_ascii=False, indent=2)

    print(f"  ✓ [{index:03d}] {title[:30]} → {note_dir} ({len(downloaded)} 文件)", flush=True)
    return {
        'index': index,
        'note_dir': note_dir,
        'downloaded_files': downloaded,
        'video_path': os.path.join(note_dir, f"视频_{note_id}.mp4") if is_video else None,
        'image_count': len(imgs),
        'has_avatar': bool(av),
        'is_video': is_video,
    }


# ── 异步下载线程池 ────────────────────────────────────────────────────────────
# 生产者-消费者：浏览器循环 put(task)，N 个 worker 并发下载。worker 只做 requests I/O，
# 不碰 js/cdp（那是浏览器循环独占）——无 CDP 争用，requests 释放 GIL，真并行。
# worker_fn(task, session) -> (media, dl_info) | None。结果按 index 排序（完成顺序不定）。

class DownloadPool:
    def __init__(self, worker_fn, n_workers=3):
        self.q = queue.Queue()
        self.results = []
        self._lock = threading.Lock()
        self._worker_fn = worker_fn
        self.n = max(1, n_workers)
        self._workers = []
        self._started = False

    def start(self):
        for _ in range(self.n):
            t = threading.Thread(target=self._loop, daemon=True)
            t.start()
            self._workers.append(t)
        self._started = True

    def _loop(self):
        sess = _new_session()  # 每个 worker 独立 session（CookieJar 非线程安全）
        while True:
            task = self.q.get()
            try:
                if task is None:           # sentinel → 退出
                    break
                res = self._worker_fn(task, sess)
                if res is not None:
                    with self._lock:
                        self.results.append(res)
            except Exception as e:
                print(f"  ❌ worker 异常: {e}", flush=True)
            finally:
                self.q.task_done()

    def put(self, task):
        if not self._started:
            self.start()
        self.q.put(task)

    def join(self):
        """发 N 个 sentinel 等所有 worker 退出。调用前所有 task 必须已 put 完。"""
        if not self._started:
            return self.results
        for _ in range(self.n):
            self.q.put(None)
        for t in self._workers:
            t.join()
        self.results.sort(key=lambda r: (r[1] or {}).get('index', 0) if isinstance(r, tuple) else 0)
        return self.results

    @property
    def pending(self):
        return self.q.qsize()


def plain_worker(task, session):
    """续传/独立场景用的 worker：只下载，不重取 URL（无浏览器）"""
    media = task['media']
    return (media, download_note_media(media, task['run_dir'], task['index'], session))


# ── 汇总 Excel ──────────────────────────────────────────────────────────────

def build_excel(notes, output_path, base_dir):
    """生成媒体汇总 Excel。notes = [(media, dl_info), ...]"""
    if not HAVE_OPENPYXL:
        print("  ⚠ openpyxl 未安装，跳过 Excel。pip install openpyxl", flush=True)
        return None
    wb = Workbook()
    ws = wb.active
    ws.title = "小红书媒体"
    headers = ['序号', '标题', '作者', '类型', '视频', '图片数', '头像', '点赞', '收藏', '评论', 'IP', '缩略图']
    widths = [6, 42, 16, 8, 14, 8, 6, 8, 8, 8, 8, 18]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True)
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    row = 2
    for idx, (media, dl) in enumerate(notes, 1):
        color = POST_COLORS[(idx - 1) % len(POST_COLORS)]
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        title = media.get('title', '')[:60]
        nick = media.get('nickname', '')
        ntype = '视频' if dl.get('is_video') else '图文'

        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=title)
        ws.cell(row=row, column=3, value=nick)
        ws.cell(row=row, column=4, value=ntype)

        # 视频列：本地路径超链接（Excel 嵌不了视频本体）
        if dl.get('video_path') and os.path.exists(dl['video_path']):
            vc = ws.cell(row=row, column=5, value='🎬 打开')
            vc.hyperlink = f"file://{dl['video_path']}"
            vc.font = Font(color="0563C1", underline="single")
        else:
            ws.cell(row=row, column=5, value='—')

        ws.cell(row=row, column=6, value=dl.get('image_count', 0))
        ws.cell(row=row, column=7, value='✓' if dl.get('has_avatar') else '—')
        ws.cell(row=row, column=8, value=str(media.get('likedCount', '')))
        ws.cell(row=row, column=9, value=str(media.get('collectedCount', '')))
        ws.cell(row=row, column=10, value=str(media.get('commentCount', '')))
        ws.cell(row=row, column=11, value=media.get('ipLocation', ''))

        # 缩略图：首张原图（OneCellAnchor，不随单元格漂移）
        if dl.get('note_dir'):
            first_img = None
            for fn in dl.get('downloaded_files', []):
                if fn.startswith('图片_'):
                    first_img = os.path.join(dl['note_dir'], fn)
                    break
                if fn.startswith('视频_') and not first_img:
                    # 视频笔记也常带一张封面图；若无图则跳过
                    pass
            if first_img and os.path.exists(first_img):
                try:
                    img = XlImage(first_img)
                    img.width = 80; img.height = 80
                    img.anchor = OneCellAnchor(
                        _from=AnchorMarker(col=11, colOff=0, row=row - 1, rowOff=0),
                        ext=XDRPositiveSize2D(pixels_to_EMU(80), pixels_to_EMU(80)),
                    )
                    ws.add_image(img)
                except Exception:
                    pass

        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).fill = fill
        ws.row_dimensions[row].height = 60
        row += 1

    wb.save(output_path)
    return output_path


# ── 入口 ────────────────────────────────────────────────────────────────────

def load_notes(path):
    if os.path.isfile(path):
        with open(path, encoding='utf-8') as f:
            return [json.load(f)]
    files = sorted(glob.glob(os.path.join(path, '*.json')))
    notes = []
    for fp in files:
        if os.path.basename(fp) == '笔记信息.json':
            continue
        try:
            with open(fp, encoding='utf-8') as f:
                notes.append(json.load(f))
        except Exception:
            pass
    return notes


def main():
    ap = argparse.ArgumentParser(description='小红书媒体下载器 + Excel')
    ap.add_argument('json_dir', help='xhs_media_extract 输出的 JSON 目录')
    ap.add_argument('output', nargs='?', default=None, help='输出 Excel 路径（可选）')
    ap.add_argument('--no-images', action='store_true', help='跳过图片（只下视频/头像）')
    args = ap.parse_args()

    base_dir = os.path.dirname(args.json_dir) if os.path.isfile(args.json_dir) else args.json_dir
    notes = load_notes(args.json_dir)
    print(f"加载 {len(notes)} 篇笔记")

    n_workers = int(os.environ.get('XHS_WORKERS', '3'))
    pool = DownloadPool(plain_worker, n_workers=n_workers)
    for i, media in enumerate(notes, 1):
        if args.no_images:
            media = dict(media); media['imageUrls'] = []
        pool.put({'media': media, 'run_dir': base_dir, 'index': i})
    print(f"🧵 {n_workers} 个 worker 并发下载中…", flush=True)
    done = pool.join()

    if HAVE_OPENPYXL:
        out = args.output or os.path.join(base_dir, f'xhs_media_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        build_excel(done, out, base_dir)
        print(f"\n✅ {len(done)} 篇 → {out}")

    videos = sum(1 for _, d in done if d.get('video_path'))
    print(f"📊 视频 {videos} 篇, 图文 {len(done) - videos} 篇")


if __name__ == '__main__':
    main()
